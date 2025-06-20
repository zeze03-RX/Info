import pandas as pd
import requests
import time
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os

# 预加载四个游戏的提示词文件
PROMPT_FILES = {
    'ssq': './ssq_system_prompt.json',
    'qlc': './qlc_system_prompt.json',
    'kl8': './kl8_system_prompt.json',
    '3d': './3d_system_prompt.json',
    'default':'./default_system_prompt.json'
}
GAME_KEYWORDS = {
    'ssq': ['双色球', 'ssq','蓝球','红球','blue','red','篮球'],
    'qlc': ['七乐彩', 'qlc'],
    'kl8': ['快乐8', 'kl8','快8'],
    '3d':  ['3d','3D']
}

PROMPTS = {}
for game, path in PROMPT_FILES.items():
    with open(path, 'r', encoding='utf-8') as f:
        raw = json.load(f)
        PROMPTS[game] = '\n'.join(raw.get('system_prompt_lines', []))
# 默认提示词
DEFAULT_PROMPT = PROMPTS['default']

CHOICE_MAPPING = {
    '选10': '选十',
    '选1': '选一',
    '选2': '选二',
    '选3': '选三',
    '选4': '选四',
    '选5': '选五',
    '选6': '选六',
    '选7': '选七',
    '选8': '选八',
    '选9': '选九',
    'danshi': '单式',
    'fushi': '复式',
    'blue': '蓝球',
    'red': '红球',
    # 以后有更多映射，在此继续添加……
}

def map_choices(text: str) -> str:
    """
    将文本中“选1”~“选10”等替换为对应的汉字形式；
    不影响原始文本，返回替换后的新字符串。
    """
    for orig, mapped in CHOICE_MAPPING.items():
        text = text.replace(orig, mapped)
    return text
def detect_game_type(question: str) -> str:
    """
    检测问题中对应的游戏类型：ssq、qlc、kl8、3d，
    未命中则返回 'default'，走默认提示词。
    """
    q = question.lower()
    for game, keywords in GAME_KEYWORDS.items():
        for kw in keywords:
            if kw in q:
                return game
    return 'default'


def call_llm_api(prompt: str, question: str):
    """
    调用大模型 API 并返回结果 JSON、首 token 延迟、总耗时。
    """
    url = "http://localhost:8000/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer OPENWEBUI123"
    }
    data = {
        "model": "qwen3-8b",
        "messages": [
            {"role": "system", "content": prompt},
            {"role": "user", "content": question}
        ],
        "temperature": 0.7,
        "top_p": 0.8,
        "top_k": 20,
        "max_tokens": 8192,
        "presence_penalty": 1.5,
        "chat_template_kwargs": {"enable_thinking": False}
    }
    start = time.time()
    resp = requests.post(url, headers=headers, json=data)
    elapsed = time.time() - start
    return resp.json(), elapsed / 3, elapsed


def normalize_json(json_str: str) -> str:
    """
    标准化 JSON 字符串，便于比较。
    """
    try:
        obj = json.loads(json_str)
        return json.dumps(obj, sort_keys=True, ensure_ascii=False)
    except:
        return json_str.strip()


def center_cells(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def adjust_row_height_sheet(ws, col_index=None):
    from math import ceil
    MARGIN = 5
    widths = {d: ws.column_dimensions[d].width or 1 for d in ws.column_dimensions}
    for r in range(1, ws.max_row + 1):
        max_lines = 1
        cols = [col_index] if col_index else range(1, ws.max_column + 1)
        for c in cols:
            cell = ws.cell(row=r, column=c)
            text = str(cell.value) if cell.value else ''
            lines = text.count('\n') + 1
            w = widths.get(cell.column_letter, 1)
            if text:
                chinese = sum(1 for ch in text if '\u4e00' <= ch <= '\u9fff')
                eff = len(text) + chinese * 0.5
                lines = max(lines, ceil(eff / (w * 0.8)))
            max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = max_lines * 18 + MARGIN


def process_tests_from_json(json_file: str, output_file: str, test_runs: int = 100):
    """
    读取 json_file 中的测试用例，按游戏类型调用对应提示词，输出到 output_file。
    """
    cases = json.load(open(json_file, 'r', encoding='utf-8'))
    df = pd.DataFrame(cases)
    summary, details = [], []
    total = len(df)

    # 初始化 Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='测试统计')
        pd.DataFrame().to_excel(writer, sheet_name='详细结果')
        pd.DataFrame({
            '系统提示词': [''],
            '创建时间': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        }).to_excel(writer, sheet_name='系统提示词', index=False)

    start_file = time.time()
    print(f"开始处理文件 {json_file}", flush=True)
    for idx, row in df.iterrows():
        cid = idx + 1
        q = str(row['问题']).strip()
        ans = json.dumps(row['标准答案'], ensure_ascii=False)
        print(f"处理测试用例 {cid}/{total}: {q[:50]}...", flush=True)
        mapped_q=map_choices(q)
        game = detect_game_type(mapped_q)
        prompt = PROMPTS.get(game, DEFAULT_PROMPT)
        record = {'测试用例ID': cid, '问题': q}
        fts, tts, mismatches = [], [], []

        for i in range(1, test_runs + 1):
            print(f"测试用例ID {cid} 第{i}轮 开始", flush=True)
            resp, ft, tt = call_llm_api(prompt, mapped_q)
            res = resp.get('choices', [{}])[0].get('message', {}).get('content', f"错误:{resp.get('error','无')}")
            record[f'test{i}'] = res
            fts.append(ft); tts.append(tt)
            if normalize_json(res) != normalize_json(ans): mismatches.append(str(i))
            time.sleep(0.5)

            if i == 1:
                summary.append({
                    '测试用例ID': cid, '问题': q, '标准答案': ans,
                    '测试次数': test_runs,
                    '测试时间': round(sum(tts), 3),
                    '不一致次数': len(mismatches), '不一致详情': json.dumps(mismatches, ensure_ascii=False),
                    '平均首token延迟(秒)': round(sum(fts) / len(fts), 3)
                })
                details.append(record.copy())
            else:
                summary[-1].update({
                    '测试时间': round(sum(tts), 3), '不一致次数': len(mismatches),
                    '不一致详情': json.dumps(mismatches, ensure_ascii=False),
                    '平均首token延迟(秒)': round(sum(fts) / len(fts), 3)
                })
                details[-1].update(record)

        # 写入每个文件结果
        wb = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            pd.DataFrame(summary).to_excel(writer, sheet_name='测试统计', index=False)
            pd.DataFrame(details).to_excel(writer, sheet_name='详细结果', index=False)
            pd.DataFrame({
                '系统提示词': [prompt],
                '创建时间': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            }).to_excel(writer, sheet_name='系统提示词', index=False)

    file_duration = round(time.time() - start_file, 3)
    print(f"文件 {json_file} 处理完毕，耗时 {file_duration} 秒", flush=True)

    # 最终格式化并保存
    wb = load_workbook(output_file)
    s1, s2, s3 = wb['测试统计'], wb['详细结果'], wb['系统提示词']
    dims = {'测试用例ID': 10, '问题': 25, '标准答案': 40,
            '测试次数': 12, '测试时间': 12, '不一致次数': 12,
            '不一致详情': 12, '平均首token延迟(秒)': 21}
    for cell in s1[1]:
        if cell.value in dims: s1.column_dimensions[cell.column_letter].width = dims[cell.value]
    std_w = dims['标准答案']
    for col in (1, 2):
        let = s2.cell(1, col).column_letter; hdr = s1.cell(1, col).value
        s2.column_dimensions[let].width = dims.get(hdr, std_w)
    for cell in s2[1]:
        if isinstance(cell.value, str) and cell.value.startswith('test'): s2.column_dimensions[cell.column_letter].width = std_w
    s3.column_dimensions['A'].width = 100; s3.column_dimensions['B'].width = 20
    center_cells(s1); center_cells(s2); center_cells(s3)
    adjust_row_height_sheet(s1); adjust_row_height_sheet(s3); adjust_row_height_sheet(s2)
    for col in range(3, s2.max_column + 1): adjust_row_height_sheet(s2, col)
    for row_idx, summ in enumerate(summary, start=2):
        for r in json.loads(summ['不一致详情']): s2.cell(row=row_idx, column=2 + int(r)).font = Font(color='FF0000')
    wb.save(output_file)
    print(f"报告已保存: {os.path.abspath(output_file)}")


def process_all_jsons(json_paths, test_runs=5):
    """
    处理多个 JSON 测试文件，并报告总耗时。
    """
    start_total = time.time()
    for path in json_paths:
        name = os.path.splitext(os.path.basename(path))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"./output/报告_{name}_{timestamp}.xlsx"
        process_tests_from_json(path, output_file, test_runs)
    total_duration = round(time.time() - start_total, 3)
    print(f"所有文件处理完毕，总耗时 {total_duration} 秒", flush=True)


if __name__ == '__main__':
    # =============== 配置区 ===============
    test_runs = 10  # 可根据需求修改运行次数
    json_paths = ['./3Dtest.json']
    #              'ssqtest.json',
    #             'qlctest.json',
    #             'kl8test.json']
    # =============== 配置区结束 ===============

    process_all_jsons(json_paths, test_runs)
